"""
Flask Application for Employee Information Management and Salary Report Export
Features: Vietnamese PDF, Email sending, Bulk download, Email status tracking
"""
from flask import Flask, render_template, request, jsonify, send_file, session
import pandas as pd
import os
import io
import zipfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)
app.secret_key = 'salary_report_secret_key_2024'

# Register Vietnamese font for PDF
FONT_PATH = os.path.join(os.path.dirname(__file__), 'fonts', 'DejaVuSans.ttf')
VIETNAMESE_FONT_AVAILABLE = False

try:
    if os.path.exists(FONT_PATH):
        pdfmetrics.registerFont(TTFont('DejaVuSans', FONT_PATH))
        VIETNAMESE_FONT_AVAILABLE = True
        print(f"Vietnamese font loaded: {FONT_PATH}")
except Exception as e:
    print(f"Could not load Vietnamese font: {e}")

# Global storage for uploaded data
data_store = {
    'thong_tin': None,
    'luong': None,
    'columns_thong_tin': [],
    'columns_luong': [],
    'employees_list': [],
    'email_status': {}  # Track email status per employee: {index: {'sent': bool, 'success': bool, 'message': str, 'time': str}}
}

# Email configuration (can be overridden via environment variables)
EMAIL_CONFIG = {
    'smtp_server': os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
    'smtp_port': int(os.environ.get('SMTP_PORT', 587)),
    'sender_email': os.environ.get('SENDER_EMAIL', ''),
    'sender_password': os.environ.get('SENDER_PASSWORD', ''),  # For Gmail, use App Password
}

# Vietnamese to ASCII mapping for fallback
VIETNAMESE_MAP = {
    'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
    'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
    'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
    'đ': 'd',
    'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
    'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
    'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
    'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
    'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
    'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
    'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
    'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
    'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
    'À': 'A', 'Á': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
    'Ă': 'A', 'Ằ': 'A', 'Ắ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
    'Â': 'A', 'Ầ': 'A', 'Ấ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
    'Đ': 'D',
    'È': 'E', 'É': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
    'Ê': 'E', 'Ề': 'E', 'Ế': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
    'Ì': 'I', 'Í': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
    'Ò': 'O', 'Ó': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
    'Ô': 'O', 'Ồ': 'O', 'Ố': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
    'Ơ': 'O', 'Ờ': 'O', 'Ớ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
    'Ù': 'U', 'Ú': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
    'Ư': 'U', 'Ừ': 'U', 'Ứ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
    'Ỳ': 'Y', 'Ý': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y',
}


def remove_accents(text):
    """Convert Vietnamese text to ASCII for PDF fallback"""
    result = ''
    for char in str(text):
        result += VIETNAMESE_MAP.get(char, char)
    return result


def is_valid_column(col_name):
    """Check if column name is valid"""
    col_str = str(col_name).strip()
    if col_str.startswith('Unnamed') or col_str.startswith('Col_') or col_str.startswith('_'):
        return False
    if col_str == 'nan' or col_str == 'NaN' or pd.isna(col_name) or col_str == '':
        return False
    try:
        float(col_str)
        return False
    except:
        pass
    return True


def find_column_by_keywords(df, keywords):
    """Find column in dataframe that contains any of the keywords"""
    for col in df.columns:
        col_lower = str(col).lower()
        for keyword in keywords:
            if keyword.lower() in col_lower:
                return col
    return None


def get_value_from_df(df, row_data, keywords):
    """Get value from row data matching column keywords"""
    col = find_column_by_keywords(df, keywords)
    if col and col in df.columns:
        val = row_data[col] if col in row_data.index else None
        if pd.notna(val):
            try:
                return float(val)
            except:
                return val
    return 0


def format_number(val):
    """Format number with thousand separator"""
    if val is None or val == 0:
        return ""
    try:
        num = float(val)
        if num == int(num):
            return f"{int(num):,}".replace(",", ".")
        return f"{num:,.0f}".replace(",", ".")
    except:
        return str(val)


def clean_dataframe(df, sheet_name):
    """Clean and process the dataframe from Excel"""
    if 'thông tin' in sheet_name.lower() or 'thong tin' in sheet_name.lower():
        # Find header rows
        header_row = None
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            row_str = str(row.values)
            if 'Họ tên' in row_str or 'STT' in row_str:
                header_row = i
                break
        
        if header_row is not None:
            main_headers = df.iloc[header_row].tolist()
            sub_headers = df.iloc[header_row + 1].tolist() if header_row + 1 < len(df) else [None] * len(main_headers)
            
            combined_headers = []
            last_valid_main = ''
            for i, (main, sub) in enumerate(zip(main_headers, sub_headers)):
                main_str = str(main).strip() if pd.notna(main) and not str(main).startswith('Unnamed') else ''
                sub_str = str(sub).strip() if pd.notna(sub) and not str(sub).startswith('Unnamed') else ''
                
                try:
                    if main_str and float(main_str):
                        main_str = ''
                except:
                    pass
                try:
                    if sub_str and float(sub_str):
                        sub_str = ''
                except:
                    pass
                
                if main_str:
                    last_valid_main = main_str
                
                if sub_str and main_str:
                    combined = f"{main_str} - {sub_str}"
                elif sub_str:
                    combined = f"{last_valid_main} - {sub_str}" if last_valid_main else sub_str
                elif main_str:
                    combined = main_str
                else:
                    combined = f'_Col_{i}'
                
                combined_headers.append(combined)
            
            data_start = header_row + 2
            if data_start < len(df):
                check_row = df.iloc[data_start]
                first_vals = [str(v) for v in check_row.head(3).values if pd.notna(v)]
                if first_vals and all(v.replace('.', '').isdigit() for v in first_vals):
                    data_start += 1
            
            df = df.iloc[data_start:]
            df.columns = combined_headers
        
        cols_to_keep = [col for col in df.columns if not str(col).startswith('_')]
        df = df[cols_to_keep]
        df = df.dropna(how='all')
        
        stt_col = None
        for col in df.columns:
            if 'STT' in str(col).upper():
                stt_col = col
                break
        
        if stt_col:
            df[stt_col] = pd.to_numeric(df[stt_col], errors='coerce')
            df = df[df[stt_col].notna() & (df[stt_col] >= 1)]
        
        df = df.reset_index(drop=True)
    
    elif 'lương' in sheet_name.lower() or 'luong' in sheet_name.lower():
        header_row = None
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            row_str = str(row.values).upper()
            if 'HỌ TÊN' in row_str or 'HO TEN' in row_str or 'NHÂN VIÊN' in row_str:
                header_row = i
                break
        
        if header_row is not None:
            main_headers = df.iloc[header_row].tolist()
            sub_headers = df.iloc[header_row + 1].tolist() if header_row + 1 < len(df) else [None] * len(main_headers)
            
            combined_headers = []
            last_valid_main = ''
            for i, (main, sub) in enumerate(zip(main_headers, sub_headers)):
                main_str = str(main).strip() if pd.notna(main) and not str(main).startswith('Unnamed') else ''
                sub_str = str(sub).strip() if pd.notna(sub) and not str(sub).startswith('Unnamed') else ''
                
                try:
                    if main_str and float(main_str):
                        main_str = ''
                except:
                    pass
                try:
                    if sub_str and float(sub_str):
                        sub_str = ''
                except:
                    pass
                
                if main_str:
                    last_valid_main = main_str
                
                if sub_str and main_str:
                    combined = f"{main_str} - {sub_str}"
                elif sub_str:
                    combined = f"{last_valid_main} - {sub_str}" if last_valid_main else sub_str
                elif main_str:
                    combined = main_str
                else:
                    combined = f'_Col_{i}'
                
                combined_headers.append(combined)
            
            data_start = header_row + 2
            if data_start < len(df):
                check_row = df.iloc[data_start]
                first_vals = [str(v) for v in check_row.head(3).values if pd.notna(v)]
                if first_vals and all(v.replace('.', '').isdigit() for v in first_vals):
                    data_start += 1
            
            df = df.iloc[data_start:]
            df.columns = combined_headers
        
        cols_to_keep = [col for col in df.columns if not str(col).startswith('_')]
        df = df[cols_to_keep]
        df = df.dropna(how='all')
        
        stt_col = None
        for col in df.columns:
            if 'STT' in str(col).upper():
                stt_col = col
                break
        
        if stt_col:
            df[stt_col] = pd.to_numeric(df[stt_col], errors='coerce')
            df = df[df[stt_col].notna() & (df[stt_col] >= 1)]
        
        df = df.reset_index(drop=True)
    
    return df


def find_employee_name_column(df):
    """Find the column that contains employee names"""
    for col in df.columns:
        col_lower = str(col).lower()
        if 'họ tên' in col_lower or 'ho ten' in col_lower or 'tên nhân viên' in col_lower:
            return col
    return None


def find_employee_email_column(df):
    """Find the column that contains employee email"""
    for col in df.columns:
        col_lower = str(col).lower()
        if 'email' in col_lower or 'mail' in col_lower or 'e-mail' in col_lower:
            return col
    return None


def get_employees_list(df):
    """Get list of employees from dataframe"""
    name_col = find_employee_name_column(df)
    email_col = find_employee_email_column(df)
    
    if name_col is None:
        return []
    
    employees = []
    for idx, row in df.iterrows():
        name = row[name_col]
        email = row[email_col] if email_col and email_col in row.index else None
        if pd.notna(name) and str(name).strip():
            emp_data = {
                'index': int(idx),
                'name': str(name).strip()
            }
            if email and pd.notna(email) and '@' in str(email):
                emp_data['email'] = str(email).strip()
            employees.append(emp_data)
    return employees


def get_salary_slip_data(employee_info, salary_data, df_info, df_luong):
    """Extract salary slip data from employee info and salary data"""
    data = {
        'ho_ten': '',
        'luong_thoa_thuan': 0,
        'luong_dong': 0,
        'so_tai_khoan': '',
        'ten_ngan_hang': '',
        'luong_thuc_te': 0,
        'bhxh': 0,
        'doan_phi': 0,
        'thue_tncn': 0,
    }
    
    name_col = find_employee_name_column(df_info)
    if name_col and name_col in employee_info.index:
        val = employee_info[name_col]
        if pd.notna(val):
            data['ho_ten'] = str(val)
    
    for col in employee_info.index:
        val = employee_info[col]
        col_lower = str(col).lower()
        if pd.notna(val):
            if 'số tài khoản' in col_lower and 'ngân hàng' in col_lower:
                data['so_tai_khoan'] = str(val)
            elif 'số tài khoản' in col_lower and not data['so_tai_khoan']:
                data['so_tai_khoan'] = str(val)
            
            if 'tại ngân hàng' in col_lower or ('ngân hàng' in col_lower and 'chi nhánh' in col_lower):
                data['ten_ngan_hang'] = str(val)
            elif 'ngân hàng' in col_lower and 'số' not in col_lower and not data['ten_ngan_hang']:
                data['ten_ngan_hang'] = str(val)
    
    if salary_data is not None:
        for col in salary_data.index:
            val = salary_data[col]
            col_str = str(col)
            col_lower = col_str.lower()
            if pd.notna(val):
                try:
                    num_val = float(val)
                except:
                    num_val = 0
                
                if ('thuế tncn' in col_lower and 'tổng thu nhập' in col_lower and 
                    'chưa' not in col_lower and 'chịu' not in col_lower and 
                    'tính' not in col_lower and 'bao gồm' not in col_lower):
                    data['luong_thoa_thuan'] = num_val
                    data['luong_thuc_te'] = num_val
                
                if 'lương cơ bản' in col_lower:
                    data['luong_dong'] = num_val
                
                if 'người lao động phải nộp' in col_lower and 'tổng cộng' in col_lower:
                    data['bhxh'] = num_val
                elif 'nld phải nộp' in col_lower and 'tổng cộng' in col_lower:
                    data['bhxh'] = num_val
                
                if 'kinh phí công đoàn' in col_lower and 'phí đoàn viên' in col_lower:
                    data['doan_phi'] = num_val
                elif 'phí đoàn viên' in col_lower and data['doan_phi'] == 0:
                    data['doan_phi'] = num_val
                
                if 'thuế tncn phải nộp' in col_lower and 'tr' not in col_lower and '%' not in col_lower:
                    data['thue_tncn'] = num_val
    
    data['tong_khoan_tru'] = data['bhxh'] + data['doan_phi'] + data['thue_tncn']
    data['tong_thu_nhap'] = data['luong_thuc_te']
    data['luong_thuc_nhan'] = data['tong_thu_nhap'] - data['tong_khoan_tru']
    
    return data


def generate_excel_salary_slip(employee_index, month, year):
    """Generate Excel salary slip and return as bytes"""
    df_info = data_store['thong_tin']
    df_luong = data_store['luong']
    
    if employee_index >= len(df_info):
        return None, None
    
    employee = df_info.iloc[employee_index]
    name_col = find_employee_name_column(df_info)
    employee_name = employee[name_col] if name_col else f'NhanVien_{employee_index}'
    
    salary_row = None
    if df_luong is not None and name_col:
        luong_name_col = find_employee_name_column(df_luong)
        if luong_name_col:
            salary_match = df_luong[df_luong[luong_name_col].astype(str).str.lower().str.strip() == str(employee_name).lower().strip()]
            if len(salary_match) > 0:
                salary_row = salary_match.iloc[0]
    
    slip_data = get_salary_slip_data(employee, salary_row, df_info, df_luong)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Phiếu Lương"
    
    title_font = Font(bold=True, size=16, color="FF0000")
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    
    ws.merge_cells('A1:E1')
    ws['A1'] = f"PHIẾU LƯƠNG THÁNG {month} NĂM {year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    info_rows = [
        ('Họ tên:', slip_data['ho_ten'], 'Ngày công chuẩn:', ''),
        ('Lương thỏa thuận:', format_number(slip_data['luong_thoa_thuan']), 'Ngày công thực tế:', ''),
        ('% Lương Thử việc:', '', 'Nghỉ phép:', ''),
        ('Lương đóng:', format_number(slip_data['luong_dong']), 'Tổng giờ tăng ca:', ''),
        ('Số tài khoản:', slip_data['so_tai_khoan'], 'Tên ngân hàng:', slip_data['ten_ngan_hang']),
    ]
    
    for info in info_rows:
        ws[f'A{row}'] = info[0]
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = yellow_fill
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = info[1]
        ws[f'B{row}'].border = border
        
        ws[f'C{row}'] = info[2]
        ws[f'C{row}'].font = header_font
        ws[f'C{row}'].fill = yellow_fill
        ws[f'C{row}'].border = border
        
        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'] = info[3]
        ws[f'D{row}'].border = border
        ws[f'E{row}'].border = border
        
        row += 1
    
    row += 1
    ws[f'A{row}'] = 'STT'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = orange_fill
    ws[f'A{row}'].border = border
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = 'Các Khoản Thu Nhập'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = orange_fill
    ws[f'B{row}'].border = border
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    ws[f'C{row}'].border = border
    
    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = 'Các Khoản Trừ Vào Lương'
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].fill = orange_fill
    ws[f'D{row}'].border = border
    ws[f'D{row}'].alignment = Alignment(horizontal='center')
    ws[f'E{row}'].border = border
    
    table_data = [
        ('1', 'Lương thực tế', format_number(slip_data['luong_thuc_te']), 'BHXH', format_number(slip_data['bhxh'])),
        ('2', 'Phép năm', '', 'Đoàn phí', format_number(slip_data['doan_phi'])),
        ('3', 'Lương tăng ca', '', 'Thuế Thu Nhập Cá Nhân', format_number(slip_data['thue_tncn'])),
        ('4', 'Lương bổ sung', '', 'Tạm Ứng', ''),
        ('5', 'Giữ xe', '', 'Tiền phạt', ''),
        ('6', 'Công tác phí', '', 'Khác', ''),
    ]
    
    row += 1
    for data_row in table_data:
        ws[f'A{row}'] = data_row[0]
        ws[f'A{row}'].border = border
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].fill = light_blue_fill
        
        ws[f'B{row}'] = data_row[1]
        ws[f'B{row}'].border = border
        ws[f'B{row}'].fill = light_blue_fill
        
        ws[f'C{row}'] = data_row[2]
        ws[f'C{row}'].border = border
        ws[f'C{row}'].fill = light_green_fill
        
        ws[f'D{row}'] = data_row[3]
        ws[f'D{row}'].border = border
        ws[f'D{row}'].fill = light_blue_fill
        
        ws[f'E{row}'] = data_row[4]
        ws[f'E{row}'].border = border
        ws[f'E{row}'].fill = light_green_fill
        
        row += 1
    
    ws[f'A{row}'] = ''
    ws[f'A{row}'].border = border
    
    ws[f'B{row}'] = 'Tổng Cộng Thu Nhập'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].border = border
    ws[f'B{row}'].fill = yellow_fill
    
    ws[f'C{row}'] = format_number(slip_data['tong_thu_nhap'])
    ws[f'C{row}'].border = border
    ws[f'C{row}'].fill = light_yellow_fill
    
    ws[f'D{row}'] = 'Tổng Cộng Khoản Trừ'
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].border = border
    ws[f'D{row}'].fill = yellow_fill
    
    ws[f'E{row}'] = format_number(slip_data['tong_khoan_tru'])
    ws[f'E{row}'].border = border
    ws[f'E{row}'].fill = light_yellow_fill
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'Tổng Số Tiền Lương Thực Nhận'
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
    ws[f'A{row}'].border = border
    ws[f'A{row}'].fill = yellow_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'B{row}'].border = border
    ws[f'C{row}'].border = border
    ws[f'D{row}'].border = border
    
    ws[f'E{row}'] = format_number(slip_data['luong_thuc_nhan'])
    ws[f'E{row}'].font = Font(bold=True, size=12, color="FF0000")
    ws[f'E{row}'].border = border
    ws[f'E{row}'].fill = light_yellow_fill
    
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Anh/Chị vui lòng kiểm tra lại thông tin trên phiếu lương. Mọi thắc mắc vui lòng liên hệ Phòng HCNS trong vòng"
    ws[f'A{row}'].font = Font(size=9, italic=True)
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "24 giờ (kể từ thời điểm nhận được thông báo này) để được giải quyết."
    ws[f'A{row}'].font = Font(size=9, italic=True)
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Quá thời hạn trên, thông tin trên phiếu lương sẽ được xem là chính xác và không có khiếu nại. Trân trọng cảm ơn!"
    ws[f'A{row}'].font = Font(size=9, italic=True)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    safe_name = "".join(c for c in str(employee_name) if c.isalnum() or c in (' ', '_')).strip()
    filename = f"PhieuLuong_{safe_name}_Thang{month}_{year}.xlsx"
    
    return output.getvalue(), filename


def generate_pdf_salary_slip(employee_index, month, year):
    """Generate PDF salary slip with Vietnamese support and return as bytes"""
    df_info = data_store['thong_tin']
    df_luong = data_store['luong']
    
    if employee_index >= len(df_info):
        return None, None
    
    employee = df_info.iloc[employee_index]
    name_col = find_employee_name_column(df_info)
    employee_name = employee[name_col] if name_col else f'NhanVien_{employee_index}'
    
    salary_row = None
    if df_luong is not None and name_col:
        luong_name_col = find_employee_name_column(df_luong)
        if luong_name_col:
            salary_match = df_luong[df_luong[luong_name_col].astype(str).str.lower().str.strip() == str(employee_name).lower().strip()]
            if len(salary_match) > 0:
                salary_row = salary_match.iloc[0]
    
    slip_data = get_salary_slip_data(employee, salary_row, df_info, df_luong)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Use Vietnamese font if available
    font_name = 'DejaVuSans' if VIETNAMESE_FONT_AVAILABLE else 'Helvetica'
    
    # Helper function for text - use Vietnamese if font available
    def vn_text(text):
        if VIETNAMESE_FONT_AVAILABLE:
            return str(text)
        return remove_accents(str(text))
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=16,
        alignment=1,
        spaceAfter=20,
        textColor=colors.red
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=10
    )
    
    note_style = ParagraphStyle(
        'CustomNote',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=8,
        italic=True
    )
    
    # Title
    title_text = f"PHIẾU LƯƠNG THÁNG {month} NĂM {year}" if VIETNAMESE_FONT_AVAILABLE else f"PHIEU LUONG THANG {month} NAM {year}"
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 10))
    
    # Info table
    info_data = [
        [vn_text('Họ tên:'), vn_text(slip_data['ho_ten']), vn_text('Ngày công chuẩn:'), ''],
        [vn_text('Lương thỏa thuận:'), format_number(slip_data['luong_thoa_thuan']), vn_text('Ngày công thực tế:'), ''],
        [vn_text('% Lương Thử việc:'), '', vn_text('Nghỉ phép:'), ''],
        [vn_text('Lương đóng:'), format_number(slip_data['luong_dong']), vn_text('Tổng giờ tăng ca:'), ''],
        [vn_text('Số tài khoản:'), slip_data['so_tai_khoan'], vn_text('Tên ngân hàng:'), vn_text(slip_data['ten_ngan_hang'])],
    ]
    
    info_table = Table(info_data, colWidths=[3.5*cm, 5*cm, 4*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.yellow),
        ('BACKGROUND', (2, 0), (2, -1), colors.yellow),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))
    
    # Main table
    table_header = [
        ['STT', vn_text('Các Khoản Thu Nhập'), '', vn_text('Các Khoản Trừ Vào Lương'), '']
    ]
    
    table_data = [
        ['1', vn_text('Lương thực tế'), format_number(slip_data['luong_thuc_te']), 'BHXH', format_number(slip_data['bhxh'])],
        ['2', vn_text('Phép năm'), '', vn_text('Đoàn phí'), format_number(slip_data['doan_phi'])],
        ['3', vn_text('Lương tăng ca'), '', vn_text('Thuế Thu Nhập Cá Nhân'), format_number(slip_data['thue_tncn'])],
        ['4', vn_text('Lương bổ sung'), '', vn_text('Tạm Ứng'), ''],
        ['5', vn_text('Giữ xe'), '', vn_text('Tiền phạt'), ''],
        ['6', vn_text('Công tác phí'), '', vn_text('Khác'), ''],
    ]
    
    totals_data = [
        ['', vn_text('Tổng Cộng Thu Nhập'), format_number(slip_data['tong_thu_nhap']), 
         vn_text('Tổng Cộng Khoản Trừ'), format_number(slip_data['tong_khoan_tru'])],
    ]
    
    full_table_data = table_header + table_data + totals_data
    
    main_table = Table(full_table_data, colWidths=[1.5*cm, 4.5*cm, 3.5*cm, 4.5*cm, 3.5*cm])
    main_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFC000')),
        ('SPAN', (1, 0), (2, 0)),
        ('SPAN', (3, 0), (4, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        
        ('BACKGROUND', (0, 1), (0, -2), colors.HexColor('#DAEEF3')),
        ('BACKGROUND', (1, 1), (1, -2), colors.HexColor('#DAEEF3')),
        ('BACKGROUND', (2, 1), (2, -2), colors.HexColor('#E2EFDA')),
        ('BACKGROUND', (3, 1), (3, -2), colors.HexColor('#DAEEF3')),
        ('BACKGROUND', (4, 1), (4, -2), colors.HexColor('#E2EFDA')),
        
        ('BACKGROUND', (1, -1), (1, -1), colors.yellow),
        ('BACKGROUND', (3, -1), (3, -1), colors.yellow),
        ('BACKGROUND', (2, -1), (2, -1), colors.HexColor('#FFFFCC')),
        ('BACKGROUND', (4, -1), (4, -1), colors.HexColor('#FFFFCC')),
        
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(main_table)
    elements.append(Spacer(1, 5))
    
    # Net salary row
    net_data = [[vn_text('Tổng Số Tiền Lương Thực Nhận'), format_number(slip_data['luong_thuc_nhan'])]]
    net_table = Table(net_data, colWidths=[14*cm, 3.5*cm])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.yellow),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#FFFFCC')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.red),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(net_table)
    
    # Footer note
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(vn_text("Anh/Chị vui lòng kiểm tra lại thông tin trên phiếu lương. Mọi thắc mắc vui lòng liên hệ Phòng HCNS trong vòng"), note_style))
    elements.append(Paragraph(vn_text("24 giờ (kể từ thời điểm nhận được thông báo này) để được giải quyết."), note_style))
    elements.append(Paragraph(vn_text("Quá thời hạn trên, thông tin trên phiếu lương sẽ được xem là chính xác và không có khiếu nại. Trân trọng cảm ơn!"), note_style))
    
    doc.build(elements)
    output.seek(0)
    
    safe_name = "".join(c for c in str(employee_name) if c.isalnum() or c in (' ', '_')).strip()
    filename = f"PhieuLuong_{safe_name}_Thang{month}_{year}.pdf"
    
    return output.getvalue(), filename


def send_email_with_attachment(to_email, subject, body, attachment_data, attachment_filename):
    """Send email with attachment using SMTP"""
    if not EMAIL_CONFIG['sender_email'] or not EMAIL_CONFIG['sender_password']:
        return False, "Chưa cấu hình email gửi. Vui lòng cấu hình SMTP_SERVER, SENDER_EMAIL, SENDER_PASSWORD."
    
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(attachment_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename="{attachment_filename}"')
        msg.attach(attachment)
        
        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.starttls()
        server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
        server.send_message(msg)
        server.quit()
        
        return True, "Email đã gửi thành công!"
    except Exception as e:
        return False, f"Lỗi gửi email: {str(e)}"


@app.route('/')
def index():
    """Main page"""
    columns_info = data_store['columns_thong_tin'] if data_store['thong_tin'] is not None else []
    columns_luong = data_store['columns_luong'] if data_store['luong'] is not None else []
    has_data = data_store['thong_tin'] is not None
    employees_list = data_store['employees_list']
    email_status = data_store['email_status']
    
    email_configured = bool(EMAIL_CONFIG['sender_email'] and EMAIL_CONFIG['sender_password'])
    
    return render_template('index.html', 
                         columns_info=columns_info,
                         columns_luong=columns_luong,
                         has_data=has_data,
                         employees_list=employees_list,
                         email_status=email_status,
                         email_configured=email_configured)


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Không tìm thấy file'})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Chưa chọn file'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'})
    
    try:
        xlsx = pd.ExcelFile(file)
        
        for sheet_name in xlsx.sheet_names:
            df = pd.read_excel(xlsx, sheet_name=sheet_name, header=None)
            df_cleaned = clean_dataframe(df, sheet_name)
            
            if 'thông tin' in sheet_name.lower() or 'thong tin' in sheet_name.lower():
                data_store['thong_tin'] = df_cleaned
                data_store['columns_thong_tin'] = [col for col in df_cleaned.columns if is_valid_column(col)]
                data_store['employees_list'] = get_employees_list(df_cleaned)
            elif 'lương' in sheet_name.lower() or 'luong' in sheet_name.lower():
                data_store['luong'] = df_cleaned
                data_store['columns_luong'] = [col for col in df_cleaned.columns if is_valid_column(col)]
        
        # Reset email status on new upload
        data_store['email_status'] = {}
        
        return jsonify({
            'success': True,
            'message': f'Đã upload thành công file: {file.filename}',
            'columns_info': data_store['columns_thong_tin'],
            'columns_luong': data_store['columns_luong'],
            'employees_list': data_store['employees_list'],
            'total_employees': len(data_store['employees_list'])
        })
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': f'Lỗi xử lý file: {str(e)}'})


@app.route('/get_employee/<int:employee_index>')
def get_employee(employee_index):
    """Get single employee data by index"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Chưa upload dữ liệu'})
    
    df_info = data_store['thong_tin']
    df_luong = data_store['luong']
    
    if employee_index >= len(df_info):
        return jsonify({'success': False, 'error': 'Không tìm thấy nhân viên'})
    
    employee = df_info.iloc[employee_index]
    name_col = find_employee_name_column(df_info)
    email_col = find_employee_email_column(df_info)
    
    employee_data = {
        'index': employee_index,
        'info': {},
        'email': None,
        'email_status': data_store['email_status'].get(employee_index)
    }
    
    if email_col and email_col in employee.index:
        email = employee[email_col]
        if pd.notna(email) and '@' in str(email):
            employee_data['email'] = str(email).strip()
    
    for col in df_info.columns:
        if is_valid_column(col):
            val = employee[col]
            if pd.notna(val):
                employee_data['info'][col] = str(val)
    
    if df_luong is not None and name_col:
        employee_name = employee[name_col]
        luong_name_col = find_employee_name_column(df_luong)
        if luong_name_col and pd.notna(employee_name):
            salary_match = df_luong[df_luong[luong_name_col].astype(str).str.lower().str.strip() == str(employee_name).lower().strip()]
            if len(salary_match) > 0:
                employee_data['salary'] = {}
                for col in df_luong.columns:
                    if is_valid_column(col):
                        val = salary_match.iloc[0][col]
                        if pd.notna(val):
                            employee_data['salary'][col] = str(val)
    
    return jsonify({
        'success': True,
        'result': employee_data
    })


@app.route('/search', methods=['POST'])
def search_employee():
    """Search for employee"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Chưa upload dữ liệu. Vui lòng upload file Excel trước.'})
    
    data = request.get_json()
    search_term = data.get('search_term', '').strip()
    
    if not search_term:
        return jsonify({'success': False, 'error': 'Vui lòng nhập từ khóa tìm kiếm'})
    
    df_info = data_store['thong_tin']
    df_luong = data_store['luong']
    
    mask = pd.Series([False] * len(df_info))
    
    for col in df_info.columns:
        if is_valid_column(col):
            mask = mask | df_info[col].astype(str).str.lower().str.contains(search_term.lower(), na=False)
    
    results = df_info[mask]
    
    if len(results) == 0:
        return jsonify({'success': False, 'error': 'Không tìm thấy kết quả'})
    
    results_list = []
    name_col = find_employee_name_column(df_info)
    email_col = find_employee_email_column(df_info)
    
    for idx, row in results.iterrows():
        employee_data = {
            'index': int(idx),
            'info': {},
            'email': None,
            'email_status': data_store['email_status'].get(int(idx))
        }
        
        if email_col and email_col in row.index:
            email = row[email_col]
            if pd.notna(email) and '@' in str(email):
                employee_data['email'] = str(email).strip()
        
        for col in df_info.columns:
            if is_valid_column(col):
                val = row[col]
                if pd.notna(val):
                    employee_data['info'][col] = str(val)
        
        if df_luong is not None and name_col:
            employee_name = row[name_col]
            luong_name_col = find_employee_name_column(df_luong)
            if luong_name_col and pd.notna(employee_name):
                salary_match = df_luong[df_luong[luong_name_col].astype(str).str.lower().str.strip() == str(employee_name).lower().strip()]
                if len(salary_match) > 0:
                    employee_data['salary'] = {}
                    for col in df_luong.columns:
                        if is_valid_column(col):
                            val = salary_match.iloc[0][col]
                            if pd.notna(val):
                                employee_data['salary'][col] = str(val)
        
        results_list.append(employee_data)
    
    return jsonify({
        'success': True,
        'results': results_list,
        'count': len(results_list)
    })


@app.route('/export/excel/<int:employee_index>', methods=['GET', 'POST'])
def export_excel(employee_index):
    """Export employee salary slip to Excel"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Không có dữ liệu'})
    
    try:
        now = datetime.now()
        month = request.args.get('month', now.month, type=int)
        year = request.args.get('year', now.year, type=int)
        
        file_data, filename = generate_excel_salary_slip(employee_index, month, year)
        
        if file_data is None:
            return jsonify({'success': False, 'error': 'Không tìm thấy nhân viên'})
        
        return send_file(
            io.BytesIO(file_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e)})


@app.route('/export/pdf/<int:employee_index>', methods=['GET', 'POST'])
def export_pdf(employee_index):
    """Export employee salary slip to PDF with Vietnamese support"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Không có dữ liệu'})
    
    try:
        now = datetime.now()
        month = request.args.get('month', now.month, type=int)
        year = request.args.get('year', now.year, type=int)
        
        file_data, filename = generate_pdf_salary_slip(employee_index, month, year)
        
        if file_data is None:
            return jsonify({'success': False, 'error': 'Không tìm thấy nhân viên'})
        
        return send_file(
            io.BytesIO(file_data),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e)})


@app.route('/export/bulk', methods=['POST'])
def export_bulk():
    """Export multiple salary slips as a zip file"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Không có dữ liệu'})
    
    try:
        data = request.get_json()
        indices = data.get('indices', [])  # List of employee indices, or 'all'
        file_type = data.get('file_type', 'pdf')  # 'pdf' or 'excel'
        month = data.get('month', datetime.now().month)
        year = data.get('year', datetime.now().year)
        
        df_info = data_store['thong_tin']
        
        # Handle 'all' selection
        if indices == 'all' or not indices:
            indices = list(range(len(df_info)))
        
        # Create zip file in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for idx in indices:
                if idx < len(df_info):
                    if file_type == 'excel':
                        file_data, filename = generate_excel_salary_slip(idx, month, year)
                    else:
                        file_data, filename = generate_pdf_salary_slip(idx, month, year)
                    
                    if file_data:
                        zip_file.writestr(filename, file_data)
        
        zip_buffer.seek(0)
        
        zip_filename = f"PhieuLuong_Thang{month}_{year}.zip"
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e)})


@app.route('/send_email/<int:employee_index>', methods=['POST'])
def send_email(employee_index):
    """Send salary slip email to employee"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Không có dữ liệu'})
    
    try:
        data = request.get_json()
        month = data.get('month', datetime.now().month)
        year = data.get('year', datetime.now().year)
        file_type = data.get('file_type', 'pdf')
        
        df_info = data_store['thong_tin']
        
        if employee_index >= len(df_info):
            return jsonify({'success': False, 'error': 'Không tìm thấy nhân viên'})
        
        employee = df_info.iloc[employee_index]
        name_col = find_employee_name_column(df_info)
        email_col = find_employee_email_column(df_info)
        
        if not email_col or email_col not in employee.index:
            return jsonify({'success': False, 'error': 'Không tìm thấy cột email trong dữ liệu'})
        
        to_email = employee[email_col]
        if not pd.notna(to_email) or '@' not in str(to_email):
            return jsonify({'success': False, 'error': 'Email nhân viên không hợp lệ'})
        
        to_email = str(to_email).strip()
        employee_name = employee[name_col] if name_col else f'Nhân viên {employee_index}'
        
        # Generate file
        if file_type == 'excel':
            file_data, filename = generate_excel_salary_slip(employee_index, month, year)
        else:
            file_data, filename = generate_pdf_salary_slip(employee_index, month, year)
        
        if not file_data:
            return jsonify({'success': False, 'error': 'Không thể tạo phiếu lương'})
        
        # Prepare email
        subject = f"Phiếu lương tháng {month}/{year}"
        body = f"""Xin chào {employee_name},

Phiếu lương tháng {month}/{year} được đính kèm bên dưới.

Anh/Chị vui lòng kiểm tra lại thông tin trên phiếu lương. Mọi thắc mắc vui lòng liên hệ Phòng HCNS trong vòng 24 giờ để được giải quyết.

Trân trọng!"""
        
        # Send email
        success, message = send_email_with_attachment(to_email, subject, body, file_data, filename)
        
        # Update status
        data_store['email_status'][employee_index] = {
            'sent': True,
            'success': success,
            'message': message,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'month': month,
            'year': year
        }
        
        return jsonify({
            'success': success,
            'message': message,
            'email_status': data_store['email_status'][employee_index]
        })
    
    except Exception as e:
        import traceback
        data_store['email_status'][employee_index] = {
            'sent': True,
            'success': False,
            'message': str(e),
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        return jsonify({'success': False, 'error': str(e)})


@app.route('/send_email_bulk', methods=['POST'])
def send_email_bulk():
    """Send salary slips to multiple employees"""
    if data_store['thong_tin'] is None:
        return jsonify({'success': False, 'error': 'Không có dữ liệu'})
    
    try:
        data = request.get_json()
        indices = data.get('indices', [])
        month = data.get('month', datetime.now().month)
        year = data.get('year', datetime.now().year)
        file_type = data.get('file_type', 'pdf')
        
        df_info = data_store['thong_tin']
        
        if indices == 'all' or not indices:
            indices = list(range(len(df_info)))
        
        results = []
        success_count = 0
        fail_count = 0
        
        for idx in indices:
            if idx < len(df_info):
                employee = df_info.iloc[idx]
                name_col = find_employee_name_column(df_info)
                email_col = find_employee_email_column(df_info)
                
                employee_name = employee[name_col] if name_col else f'NV {idx}'
                
                if not email_col or email_col not in employee.index:
                    fail_count += 1
                    data_store['email_status'][idx] = {
                        'sent': True, 'success': False, 
                        'message': 'Không có email', 
                        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    results.append({'index': idx, 'name': employee_name, 'success': False, 'message': 'Không có email'})
                    continue
                
                to_email = employee[email_col]
                if not pd.notna(to_email) or '@' not in str(to_email):
                    fail_count += 1
                    data_store['email_status'][idx] = {
                        'sent': True, 'success': False, 
                        'message': 'Email không hợp lệ', 
                        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    results.append({'index': idx, 'name': employee_name, 'success': False, 'message': 'Email không hợp lệ'})
                    continue
                
                to_email = str(to_email).strip()
                
                # Generate file
                if file_type == 'excel':
                    file_data, filename = generate_excel_salary_slip(idx, month, year)
                else:
                    file_data, filename = generate_pdf_salary_slip(idx, month, year)
                
                if not file_data:
                    fail_count += 1
                    results.append({'index': idx, 'name': employee_name, 'success': False, 'message': 'Không thể tạo file'})
                    continue
                
                # Send email
                subject = f"Phiếu lương tháng {month}/{year}"
                body = f"""Xin chào {employee_name},

Phiếu lương tháng {month}/{year} được đính kèm bên dưới.

Anh/Chị vui lòng kiểm tra lại thông tin trên phiếu lương. Mọi thắc mắc vui lòng liên hệ Phòng HCNS trong vòng 24 giờ để được giải quyết.

Trân trọng!"""
                
                success, message = send_email_with_attachment(to_email, subject, body, file_data, filename)
                
                data_store['email_status'][idx] = {
                    'sent': True, 'success': success, 
                    'message': message, 
                    'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'month': month, 'year': year
                }
                
                if success:
                    success_count += 1
                else:
                    fail_count += 1
                
                results.append({'index': idx, 'name': employee_name, 'success': success, 'message': message})
        
        return jsonify({
            'success': True,
            'message': f'Đã gửi {success_count} email thành công, {fail_count} thất bại',
            'results': results,
            'success_count': success_count,
            'fail_count': fail_count
        })
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e)})


@app.route('/email_status')
def get_email_status():
    """Get email status for all employees"""
    return jsonify({
        'success': True,
        'email_status': data_store['email_status']
    })


@app.route('/configure_email', methods=['POST'])
def configure_email():
    """Configure email settings (runtime configuration)"""
    try:
        data = request.get_json()
        
        if 'smtp_server' in data:
            EMAIL_CONFIG['smtp_server'] = data['smtp_server']
        if 'smtp_port' in data:
            EMAIL_CONFIG['smtp_port'] = int(data['smtp_port'])
        if 'sender_email' in data:
            EMAIL_CONFIG['sender_email'] = data['sender_email']
        if 'sender_password' in data:
            EMAIL_CONFIG['sender_password'] = data['sender_password']
        
        return jsonify({
            'success': True,
            'message': 'Cấu hình email đã được cập nhật',
            'config': {
                'smtp_server': EMAIL_CONFIG['smtp_server'],
                'smtp_port': EMAIL_CONFIG['smtp_port'],
                'sender_email': EMAIL_CONFIG['sender_email'],
                'has_password': bool(EMAIL_CONFIG['sender_password'])
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/get_columns')
def get_columns():
    """Get available columns for search"""
    return jsonify({
        'columns_info': data_store['columns_thong_tin'],
        'columns_luong': data_store['columns_luong'],
        'employees_list': data_store['employees_list']
    })


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_DEBUG', 'True').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=port)
